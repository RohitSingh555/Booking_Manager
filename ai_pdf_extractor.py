import pdfplumber
import openpyxl
import os
import logging
import ollama
import json
import re

# Setup basic configuration for logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def extract_text_from_pdf(pdf_path):
    """Extract text from PDF using pdfplumber."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page_number, page in enumerate(pdf.pages, start=1):
                page_text = page.extract_text()
                # logging.info("Extracted text from page %d of PDF: %s", page_number, page_text)
                text += page_text + "\n"
            return text
    except Exception as e:
        logging.error("Failed to process PDF %s: %s", pdf_path, str(e))
        return ""

def preprocess_text(text):
    """Preprocess extracted text to capture all relevant lines."""
    lines = text.split('\n')
    return "\n".join(lines)

def process_data_with_ollama(text):
    """Process data with Ollama AI to extract transactions."""
    data = []
    try:
        prompt = (
            "Please note: I don't want code! \n Take the data given below and give me a json which has Date, Amount(only keep integers in the amount), Description, Source (Upwork, Employer, Bank, Food, Housing, Utilities, Food, Supplies, Travel, Business Expense) and category (give the category from these 6 'Income, Expenses, Business Expenses, Uncertain Expenses, Tax Deductible Expenses, Subscriptions') analyze the data and description to give me a source of the transactions and category don't provide null, and always return json for the whole data don't skip anything. And even if all the transactions are expenses keep categorizing them."
            "{\n"
            '    "Date": "MM-dd-yyyy",\n'
            '    "Description": "Product or service bought, fetch the description of it",\n'
            '    "Amount": "Amount took to buy that resource.",\n'
            '    "Category": "Categorize the payments according to the description"\n'
            "}\n"
            f"data: {text}"
        )
        print("Loading...")
        response = ollama.generate(
            model='llama3',
            prompt=prompt
        )
        print("Loading complete.")
        logging.info("API Response: %s", response)

        # Check if response contains 'response' key and is not empty
        if 'response' not in response or not response['response']:
            raise ValueError("Empty or invalid response from Ollama AI")

        parsed_data = response['response']

        # Extract JSON from response and save it to a file
        with open('processed_files/transactions.json', 'w') as json_file:
            json.dump(parsed_data, json_file, indent=4)
            logging.info("JSON data saved to processed_files/transactions.json")

        transactions = json.loads(parsed_data)
        for transaction in transactions:
            data.append((transaction['Date'], transaction['Description'], transaction['Amount'], transaction['Category']))
    except json.JSONDecodeError as e:
        logging.error("JSON decoding failed: %s", str(e))
        # logging.error("Response content: %s", response)
    except Exception as e:
        logging.error("Failed to process data with Ollama AI: %s", str(e))
        # logging.error("Response content: %s", response)
    return data

def save_to_excel(data, file_name, sheet_name):
    """Save extracted data to an Excel file."""
    try:
        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove the default sheet

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(("Date", "Description", "Amount", "Category"))  # Only add headers if sheet is newly created
        else:
            ws = wb[sheet_name]

        for row in data:
            ws.append(row)

        # Remove duplicate rows
        remove_duplicate_rows(ws)

        wb.save(file_name)
        logging.info("Data written to %s in %s", sheet_name, file_name)
    except Exception as e:
        logging.error("Failed to save data to Excel %s: %s", file_name, str(e))

def remove_duplicate_rows(sheet):
    """Remove duplicate rows from an Excel sheet."""
    rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column)]
    unique_rows = set(rows)
    sheet.delete_rows(2, sheet.max_row)
    for row in unique_rows:
        sheet.append(row)

def identify_and_process_pdfs(directory, output_excel):
    """Identify PDF files in a directory and process them to extract transaction data."""
    pdf_files = [f for f in os.listdir(directory) if f.endswith('.pdf')]
    logging.info("Found %d PDF files in the directory.", len(pdf_files))
    
    for filename in pdf_files:
        pdf_path = os.path.join(directory, filename)
        logging.info("Processing PDF: %s", pdf_path)
        
        text = extract_text_from_pdf(pdf_path)
        preprocessed_text = preprocess_text(text)
        
        data = process_data_with_ollama(preprocessed_text)
        if not data:
            logging.warning("No data extracted from PDF: %s", filename)
        else:
            save_to_excel(data, output_excel, "Transactions")

def main():
    directory_path = "client_docs"
    output_excel = "processed_files/AI_pdf_output_data.xlsx"
    if not os.path.exists(directory_path):
        logging.warning("Directory %s does not exist. Please check the path.", directory_path)
        return
    identify_and_process_pdfs(directory_path, output_excel)

if __name__ == "__main__":
    main()
