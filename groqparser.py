import openpyxl
import json
import os
import time
from dotenv import load_dotenv
from groq import Groq

# Load environment variables
load_dotenv()
api_key = os.getenv("GROQ_API_KEY")

# Categories
categories = [
    "Income",
    "Expenses",
    "Business Expenses",
    "Tax Deductible Expenses",
    "Subscriptions",
    "Uncertain Expenses",
]

def read_excel_file(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheets_data = {}

        for sheet in wb.worksheets:
            max_row = sheet.max_row
            max_column = sheet.max_column

            date_column = None
            description_column = None
            amount_column = None

            for col_idx in range(1, max_column + 1):
                cell_value = sheet.cell(row=1, column=col_idx).value
                if cell_value:
                    cell_value_lower = cell_value.lower()
                    if any(keyword in cell_value_lower for keyword in ["date", "total"]):
                        date_column = col_idx
                    elif "description" in cell_value_lower:
                        description_column = col_idx
                    elif "amount" in cell_value_lower:
                        amount_column = col_idx

            # Skip the sheet if all required columns are not found
            if date_column is None or description_column is None or amount_column is None:
                print(f"Skipping sheet {sheet.title}: required columns not found.")
                continue

            # Extract data from the identified columns
            data = []
            for row_idx in range(2, max_row + 1):
                entry = {
                    "date": sheet.cell(row=row_idx, column=date_column).value,
                    "description": sheet.cell(row=row_idx, column=description_column).value,
                    "amount": sheet.cell(row=row_idx, column=amount_column).value
                }
                data.append(entry)

            sheets_data[sheet.title] = data

        return sheets_data
    except Exception as e:
        print(f"An error occurred while reading the Excel file: {e}")
        return None

def create_excel_file(response_data, categories, file_name):
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    sheets = {}
    for category in categories:
        if category in wb.sheetnames:
            sheets[category] = wb[category]
        else:
            sheets[category] = wb.create_sheet(title=category)
            sheets[category].append(["Date", "Amount", "Description", "Source"])

    for entry in response_data:
        # Normalize keys to lowercase
        normalized_entry = {k.lower(): v for k, v in entry.items()}
        category = normalized_entry.get("category")
        if category in sheets:
            sheets[category].append([normalized_entry.get("date"), normalized_entry.get("amount"), normalized_entry.get("description"), normalized_entry.get("source")])

    try:
        wb.save(file_name)
    except Exception as e:
        print(f"An error occurred while saving the Excel file: {e}")

def get_groq_response(categorized_data):
    try:
        client = Groq(api_key=api_key)

        # Calculate the context length of the messages
        prompt_message = f"Please note: I don't want code! {json.dumps(categorized_data)} \n Take this data and give me a json which has Date, Amount(only keep integers in the amount), Description, Source (Upwork, Employer, Bank, Food, Housing, Utilities, Food, Supplies, Travel, Business Expense) and category (give the category from these 6 'Income, Expenses, Business Expenses, Uncertain Expenses, Tax Deductible Expenses, Subscriptions') analyze the data and description to give me a source of the transactions and category don't provide null, and always return json for the whole data don't skip anything. And even if all the transactions are expenses keep categorizing them."
        context_length = len(prompt_message)
        print(f"Context length: {context_length}")

        # Define a limit for the context length (e.g., 6000 tokens, which is a common limit)
        context_limit = 6000

        # Truncate the prompt if it exceeds the limit
        if context_length > context_limit:
            print(f"Truncating context from {context_length} to {context_limit}")
            truncated_message = prompt_message[:context_limit]
        else:
            truncated_message = prompt_message

        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": truncated_message,
                }
            ],
            model="llama3-70b-8192",
        )

        response = chat_completion.choices[0].message.content
        print("Raw response from Groq:", response)
        return response
    except Exception as e:
        print(f"An error occurred while getting Groq response: {e}")
        return None

def extract_json_from_string(string):
    try:
        json_objects = []
        stack = []
        start_index = -1

        for i, char in enumerate(string):
            if char == '{':
                if not stack:
                    start_index = i
                stack.append('{')
            elif char == '}':
                if stack:
                    stack.pop()
                    if not stack:
                        json_str = string[start_index:i + 1]
                        try:
                            json_obj = json.loads(json_str)
                            json_objects.append(json_obj)
                        except json.JSONDecodeError:
                            continue

        return json_objects
    except Exception as e:
        print(f"An error occurred while extracting JSON from string: {e}")
        return None

def process_sheet(sheet_data, categories, file_name):
    try:
        batch_size = 12  # Number of rows to process in each batch
        for i in range(0, len(sheet_data), batch_size):
            batch = sheet_data[i:i + batch_size]
            response = get_groq_response(batch)

            try:
                json_objects = extract_json_from_string(response)
                if not json_objects:
                    raise ValueError("No valid JSON objects found in the response.")
            except json.JSONDecodeError as e:
                print(f"Error parsing JSON from response: {e}")
                continue

            print("Extracted JSON objects:", json_objects)
            create_excel_file(json_objects, categories, file_name)

            # Add delay between API requests
            time.sleep(2)
    except Exception as e:
        print(f"An error occurred while processing the sheet: {e}")

def main():
    try:
        directory_path = "processed_files"
        output_excel = "ollama/categorized_data.xlsx"

        if not os.path.exists(directory_path):
            print(f"Directory {directory_path} does not exist. Please check the path.")
            return

        # Process all .xlsx files in the directory
        for file_name in os.listdir(directory_path):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(directory_path, file_name)
                print(f"Processing file: {file_path}")
                sheets_data = read_excel_file(file_path)
                for sheet_name, sheet_data in sheets_data.items():
                    if sheet_data is None:
                        continue
                    print(f"Processing sheet: {sheet_name} in file: {file_name}")
                    process_sheet(sheet_data, categories, output_excel)
    except Exception as e:
        print(f"An error occurred in the main function: {e}")

if __name__ == "__main__":
    main()

               
