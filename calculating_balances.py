import openpyxl
from datetime import datetime, timedelta
import re

def read_excel_file(file_path, sheets_to_read):
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"Error: The file {file_path} was not found.")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    data = {sheet: [] for sheet in sheets_to_read}

    for sheet_name in sheets_to_read:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None or row[1] is None:
                    continue  # Skip rows with missing essential data
                entry = {
                    "date": row[0],
                    "amount": row[1],
                    "description": row[2] if row[2] else "",
                    "source": row[3] if row[3] else ""
                }
                data[sheet_name].append(entry)
        else:
            print(f"Sheet {sheet_name} not found in the file.")

    return data

def parse_date(date_str):
    for fmt in ('%m/%d/%y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%b %d, %Y'):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    raise ValueError(f"Date format for {date_str} is not recognized.")

def preprocess_amount(amount_str):
    try:
        # Convert to string if amount_str is not already a string
        if not isinstance(amount_str, str):
            amount_str = str(amount_str)
        
        # Remove any non-numeric characters using regex
        cleaned_amount = re.sub(r'[^0-9.-]', '', amount_str)
        # If there is a '-' at the end, treat it as a negative sign
        cleaned_amount = cleaned_amount.lstrip('-')
        # Convert the cleaned amount to a float
        float_amount = float(cleaned_amount)
        return float_amount
    except ValueError:
        print(f"Error converting amount: {amount_str} is not a valid number.")
        return 0

def calculate_weekly_balances(data):
    if "Income" not in data or "Expenses" not in data:
        print("Required data for calculating weekly balances is missing.")
        return []

    combined_data = data["Income"] + data["Expenses"]
    if not combined_data:
        return []

    combined_data.sort(key=lambda x: parse_date(x["date"]))

    start_date = parse_date(combined_data[0]["date"])
    end_date = parse_date(combined_data[-1]["date"])

    weekly_data = []
    current_date = start_date
    while current_date <= end_date:
        week_start = current_date
        week_end = current_date + timedelta(days=6)
        weekly_income = sum(preprocess_amount(entry["amount"]) for entry in data["Income"] if week_start <= parse_date(entry["date"]) <= week_end)
        weekly_expenses = sum(preprocess_amount(entry["amount"]) for entry in combined_data if week_start <= parse_date(entry["date"]) <= week_end and entry in data["Expenses"])

        weekly_balance = weekly_income - weekly_expenses
        weekly_data.append({
            "week_start": week_start.strftime('%m/%d/%y'),
            "income": weekly_income,
            "expenses": weekly_expenses,
            "balance": weekly_balance
        })

        current_date += timedelta(days=7)

    return weekly_data

def calculate_balance_summary(data, account_balances):
    total_income = sum(preprocess_amount(entry["amount"]) for entry in data.get("Income", []))

    total_expenses = sum(
        preprocess_amount(entry["amount"]) for sheet in data.keys() if sheet != "Income" for entry in data[sheet]
    )

    net_income = total_income - total_expenses

    # Adding user-inputted balances to the total income
    total_balance = sum(balance for balance in account_balances.values())

    available_budget = net_income + total_balance
    daily_budget = available_budget / 365
    weekly_budget = available_budget / 52
    yearly_budget = available_budget

    return {
        "Total Income": total_income,
        "Total Expenses": total_expenses,
        "Net Income": net_income,
        "Total Balance": total_balance,
        "Available Budget": available_budget,
        "Daily Budget": daily_budget,
        "Weekly Budget": weekly_budget,
        "Yearly Budget": yearly_budget
    }

def write_to_excel(file_path, weekly_balances, balance_summary, account_balances):
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    # Handle Weekly Budget sheet
    weekly_budget_sheet = wb.create_sheet(title="Weekly Budget") if "Weekly Budget" not in wb.sheetnames else wb["Weekly Budget"]
    weekly_budget_sheet.delete_rows(2, weekly_budget_sheet.max_row + 1)
    weekly_budget_sheet.append(["Week Start", "Income", "Expenses", "Balance"])
    for weekly_balance in weekly_balances:
        weekly_budget_sheet.append([
            weekly_balance["week_start"],
            weekly_balance["income"],
            weekly_balance["expenses"],
            weekly_balance["balance"]
        ])

    # Handle Balance Summary sheet
    balance_summary_sheet = wb.create_sheet(title="Balance Summary") if "Balance Summary" not in wb.sheetnames else wb["Balance Summary"]
    balance_summary_sheet.delete_rows(2, balance_summary_sheet.max_row + 1)
    for description, amount in balance_summary.items():
        balance_summary_sheet.append([description, amount])

    # Handle Balances sheet
    balances_sheet = wb.create_sheet(title="Balances") if "Balances" not in wb.sheetnames else wb["Balances"]
    balances_sheet.delete_rows(2, balances_sheet.max_row + 1)
    balances_sheet.append(["Account Type", "Amount"])
    for account_type, balance in account_balances.items():
        balances_sheet.append([account_type, balance])

    try:
        wb.save(file_path)
    except Exception as e:
        print(f"Error saving workbook: {e}")

def main():
    file_path = "ollama/categorized_data.xlsx"
    sheets_to_read = ["Income", "Expenses", "Business Expenses", "Tax Deductible Expenses", "Subscriptions", "Uncertain Expenses"]
    
    data = read_excel_file(file_path, sheets_to_read)
    if data is None:
        return

    if not data["Income"] or not any(data[sheet] for sheet in sheets_to_read if sheet != "Income"):
        print("No data found in Income or Expense sheets.")
        return

    weekly_balances = calculate_weekly_balances(data)

    # Get account balances from the user
    account_balances = {}
    try:
        num_accounts = int(input("Enter the number of accounts: "))
    except ValueError:
        print("Invalid number entered. Please enter a valid integer for the number of accounts.")
        return

    for _ in range(num_accounts):
        account_type = input("Enter Account Type: ").strip()
        while True:
            balance_input = input(f"Enter Balance for {account_type}: ")
            try:
                balance = float(balance_input)
                break
            except ValueError:
                print(f"Error: '{balance_input}' is not a valid float. Please enter a numeric balance.")

        account_balances[account_type] = balance

    balance_summary = calculate_balance_summary(data, account_balances)
    write_to_excel(file_path, weekly_balances, balance_summary, account_balances)
    print("Weekly Budget, Balance Summary, and Account Balances have been successfully written to the file.")

if __name__ == "__main__":
    main()
