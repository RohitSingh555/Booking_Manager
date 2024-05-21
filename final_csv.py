import openpyxl

def classify_expense(description):
    if not isinstance(description, str):
        # If description is not a string, convert it to string
        description = str(description)

    # Example classification logic
    if "paypal" in description.lower():
        return "Business"
    elif "withdrawl" in description.lower():
        return "Expenses"
    elif "office" in description.lower() or "meeting" in description.lower():
        return "Business Expense"
    else:
        return "Uncategorized"

def parse_excel(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

    parsed_data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []
        
        # Extract column names from the first row
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        column_names = [str(cell).lower() if cell is not None else "" for cell in first_row]

        # Find the indices of columns with desired names or containing desired names
        date_index = column_names.index('date') if 'date' in column_names else None
        amount_index = column_names.index('amount') if 'amount' in column_names else None
        description_index = column_names.index('description') if 'description' in column_names else None

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip first row (header)
            # Extract values based on column indices
            date = row[date_index] if date_index is not None else None
            description = row[description_index] if description_index is not None else None
            amount = row[amount_index] if amount_index is not None else None

            # Classify expense based on description
            category = classify_expense(description)

            # Append the row data along with category
            data.append((date, description, amount, category))
        
        # Append the "Category" column header to the first row
        data.insert(0, ("Date", "Description", "Amount", "Category"))
        parsed_data[sheet_name] = data

    return parsed_data


def write_to_excel(parsed_data, output_file):
    workbook = openpyxl.Workbook()
    for sheet_name, data in parsed_data.items():
        sheet = workbook.create_sheet(title=sheet_name)
        for entry in data:
            sheet.append(entry)
    workbook.remove(workbook["Sheet"])  # Remove default sheet
    workbook.save(output_file)

if __name__ == "__main__":
    input_file = "processed_files/test.xlsx"
    output_file = "parsed_files/parsed_data.xlsx"
    parsed_data = parse_excel(input_file)
    if parsed_data is not None:
        write_to_excel(parsed_data, output_file)
        print(f"Parsed data saved to {output_file}")
