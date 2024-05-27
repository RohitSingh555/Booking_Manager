import pdfplumber
import openpyxl
import os
import logging
import re

# Setup basic configuration for logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def extract_data_from_paypal(pdf_path):
    data = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                text = page.extract_text()
                logging.info("Extracted text from page %d of PayPal PDF: %s", page_number, text)
                
                if not text:
                    logging.warning("No text found on page %d of %s", page_number, pdf_path)
                    continue

                # Check if "Transaction History - USD" section is present
                if "Transaction History - USD" in text:
                    transaction_text = text.split("Transaction History - USD")[1]
                    logging.info("Transaction text extracted: %s", transaction_text)

                    # Extract transactions
                    transactions = re.findall(r'(\d{2}/\d{2}/\d{2,4})\s+(.+?)\s+([-.\d,]+)\s+([-.\d,]+)\s+([-.\d,]+)', transaction_text)
                    for transaction in transactions:
                        date, description, gross, fee, net = transaction
                        description = description.split('ID:')[0].strip()  # Clean up description
                        data.append((date, description, net, "PayPal"))
                else:
                    # Preserve the previous data extraction logic for PayPal PDF
                    lines = text.split("\n")
                    for line in lines:
                        if re.match(r'\d{2}/\d{2}/\d{4}', line):
                            parts = line.split()
                            date = parts[0]
                            description = " ".join(parts[1:-2])
                            total = parts[-1]
                            data.append((date, description, total, "PayPal"))

    except Exception as e:
        logging.error("Failed to process PDF %s: %s", pdf_path, str(e))
    return data

def extract_data_from_ebay(pdf_path):
    data = set()  # Using a set to store unique data
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                text = page.extract_text()
                # logging.info("Extracted text from page %d of eBay PDF: %s", page_number, text)
                
                if not text:
                    logging.warning("No text found on page %d of %s", page_number, pdf_path)
                    continue

                orders = text.split("Order date:")
                for order_text in orders[1:]:  # Skip first element as it's not an order
                    order_lines = order_text.split("\n")
                    date = order_lines[0].split("•")[0].strip()
                    total_line = next((line for line in order_lines if "Order total:" in line), "")
                    total = total_line.split("Order total:")[1].replace("US $", "").strip().split("•")[0] if total_line else ""
                    description = " ".join(order_lines[2:]).strip()
                    data.add((date, description, f"${total}", "eBay"))  # Adding data to the set
    except Exception as e:
        logging.error("Failed to process PDF %s: %s", pdf_path, str(e))
    return list(data)  # Converting set back to list before returning

def save_to_excel(data, file_name, sheet_name):
    try:
        if os.path.exists(file_name):
            wb = openpyxl.load_workbook(file_name)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove the default sheet

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(("Date", "Description", "Amount", "Category"))  # Only add headers if sheet is newly created
        else:
            ws = wb[sheet_name]

        for row in data:
            ws.append(row)

        # Remove duplicate rows
        remove_duplicate_rows(ws)

        wb.save(file_name)
        logging.info("Data written to %s in %s", sheet_name, file_name)
    except Exception as e:
        logging.error("Failed to save data to Excel %s: %s", file_name, str(e))

def remove_duplicate_rows(sheet):
    # Convert rows to tuples for comparison
    rows = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column)]
    unique_rows = set(rows)
    
    # Clear existing data
    sheet.delete_rows(2, sheet.max_row)
    
    # Write unique rows back to the sheet
    for row in unique_rows:
        sheet.append(row)

def identify_and_process_pdfs(directory, output_excel):
    pdf_files = [f for f in os.listdir(directory) if f.endswith('.pdf')]
    logging.info("Found %d PDF files in the directory.", len(pdf_files))
    
    for filename in pdf_files:
        pdf_path = os.path.join(directory, filename)
        logging.info("Processing PDF: %s", pdf_path)
        
        if "ebay" in filename.lower():
            logging.info("Processing eBay PDF: %s", filename)
            data = extract_data_from_ebay(pdf_path)
            save_to_excel(data, output_excel, "eBay")
        elif filename.endswith("PDF"):
            logging.info("Skipping PDF: %s", filename)
        else:
            logging.info("Processing PayPal PDF: %s", filename)
            data = extract_data_from_paypal(pdf_path)
            if filename.lower() == "creed.pdf":
                print("Extracted data from creed.pdf:")
            if not data:
                logging.warning("No data extracted from PayPal PDF: %s", filename)
            else:
                save_to_excel(data, output_excel, "PayPal")

def main():
    directory_path = "client_docs"
    output_excel = "processed_files/pdf_output_data.xlsx"
    if not os.path.exists(directory_path):
        logging.warning("Directory %s does not exist. Please check the path.", directory_path)
        return
    identify_and_process_pdfs(directory_path, output_excel)

if __name__ == "__main__":
    main()
